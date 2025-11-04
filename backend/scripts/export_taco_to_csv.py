import argparse
import os
import sys
import csv
from typing import List, Optional

from openpyxl import load_workbook

# Garantir que possamos importar utilitários do ETL existente
REPO_BACKEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if REPO_BACKEND_DIR not in sys.path:
    sys.path.append(REPO_BACKEND_DIR)

try:
    # Reutiliza mapeamento de headers e parsing de números do ETL real
    from app.services.etl_taco import _map_headers, _parse_float
except Exception:
    _map_headers = None  # será definido fallback abaixo
    _parse_float = None

def _clean_text(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    import unicodedata
    s2 = unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("ASCII")
    return s2.strip().lower()

def _fallback_map_headers(headers: List[str]):
    """Fallback local de mapeamento caso import falhe."""
    normalized = [_clean_text(h) for h in headers]

    def find_idx(*keywords):
        for i, h in enumerate(normalized):
            if h is None:
                continue
            if all(k in h for k in keywords):
                return i
        return None

    return {
        "name_pt": find_idx("alimento") or find_idx("descricao") or find_idx("nome"),
        "category_pt": find_idx("grupo") or find_idx("categoria"),
        "energy_kcal_100g": find_idx("energia", "kcal") or find_idx("kcal"),
        "energy_kj_100g": find_idx("energia", "kj") or find_idx("kj"),
        # carboidratos: cobrir variações
        "carbohydrates_100g": (
            find_idx("carbohydrates_100g")
            or find_idx("carbo", "g")
            or find_idx("carboidratos")
            or find_idx("carboidrato")
            or find_idx("idrato")
            or find_idx("carbo")
            or find_idx("carboidratos", "total")
            or find_idx("carboidrato", "total")
            or find_idx("carboidratos", "disponiveis")
            or find_idx("disponiveis", "carbo")
        ),
        # proteínas: cobrir plural e abreviações
        "proteins_100g": (
            find_idx("proteins_100g")
            or find_idx("proteina")
            or find_idx("proteinas")
            or find_idx("proteina", "g")
            or find_idx("prot")
        ),
        # gorduras/lipídeos
        "fat_100g": find_idx("lipidio") or find_idx("lipideos") or find_idx("gordura"),
        "fiber_100g": find_idx("fibra") or find_idx("alimentar"),
        "sugars_100g": find_idx("acucar") or find_idx("açucar") or find_idx("sacarose"),
        "sodium_mg_100g": find_idx("sodio") or find_idx("sódio"),
        "glycemic_index": find_idx("indice", "glicemico") or find_idx("ig"),
    }

def _fallback_parse_float(val: Optional[str]) -> Optional[float]:
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        v = str(val).replace(",", ".").strip()
        if v == "" or v.lower() in ("na", "nd", "n/a", "-"):
            return None
        return float(v)
    except Exception:
        return None

if _map_headers is None:
    _map_headers = _fallback_map_headers
if _parse_float is None:
    _parse_float = _fallback_parse_float


def export_csv_from_xlsx(xlsx_path: str, csv_path: str) -> int:
    """Exporta TACO XLSX para CSV leve usando a mesma lógica de detecção do ETL.

    - Detecta cabeçalhos pela primeira linha com >=4 células preenchidas (até linhas 1..10)
    - Usa mapeamento robusto de colunas
    - Converte números com vírgula/ponto
    - Ignora linhas sem nome
    Retorna a quantidade de linhas escritas.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Carregar primeiras linhas para permitir detecção de cabeçalho multi-linha (nome + unidade)
    rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=80, values_only=True)]
    if not rows:
        raise ValueError("Planilha vazia ou não lida")

    best_idx = None
    best_headers: List[str] = []
    best_score = -1

    def normalize_row(row_vals: List[Optional[str]]) -> List[str]:
        return [str(c) if c is not None else "" for c in row_vals]

    # Avaliar candidatos a cabeçalho nas primeiras 60 linhas
    for i in range(min(len(rows) - 1, 60)):
        header_candidates = normalize_row(rows[i])
        unit_candidates = normalize_row(rows[i + 1]) if i + 1 < len(rows) else []

        # Construir cabeçalho combinado: "proteina (g)", "energia (kcal)", etc.
        combined = []
        for j in range(len(header_candidates)):
            h = header_candidates[j] or ""
            u = unit_candidates[j] if j < len(unit_candidates) else ""
            combined.append((h + " " + u).strip())

        # Verificar se há indícios de que esta linha é cabeçalho (nome do alimento presente)
        norm = [_clean_text(c) for c in combined]
        if not any(h and ("descricao" in h or "alimento" in h or "nome" in h) for h in norm):
            continue

        # Calcular score pelo número de campos mapeados
        test_map = _map_headers(combined)
        score = sum(1 for k in [
            "name_pt","energy_kcal_100g","energy_kj_100g","carbohydrates_100g",
            "proteins_100g","fat_100g","fiber_100g","sugars_100g","sodium_mg_100g"
        ] if test_map.get(k) is not None)

        if score > best_score:
            best_score = score
            best_idx = i
            best_headers = combined

    if best_idx is None or not best_headers:
        raise ValueError("Não foi possível identificar cabeçalhos na planilha TACO")

    header_row = rows[best_idx]
    headers = best_headers

    col_map = _map_headers(headers)
    if col_map.get("name_pt") is None:
        raise ValueError("Coluna de nome do alimento não detectada nos cabeçalhos")

    # Debug: mostrar mapeamento de colunas e cabeçalhos detectados
    try:
        debug_selected = {k: (v, headers[v] if v is not None and v < len(headers) else None) for k, v in col_map.items()}
        print("Mapeamento de colunas detectado:")
        for k in [
            "name_pt","category_pt","energy_kcal_100g","energy_kj_100g",
            "carbohydrates_100g","proteins_100g","fat_100g","fiber_100g",
            "sugars_100g","sodium_mg_100g","glycemic_index"
        ]:
            idx, hdr = debug_selected.get(k, (None, None))
            print(f"  {k}: idx={idx} header='{hdr}'")
        print("Cabeçalhos combinados (amostra):")
        for i, h in enumerate(headers[:30]):
            print(f"  [{i}] '{h}'")
    except Exception:
        pass

    fieldnames = [
        "name_pt",
        "category_pt",
        "energy_kcal_100g",
        "energy_kj_100g",
        "carbohydrates_100g",
        "proteins_100g",
        "fat_100g",
        "fiber_100g",
        "sugars_100g",
        "sodium_mg_100g",
        "glycemic_index",
    ]

    written = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        data_started = False
        row_idx = -1
        for row in ws.iter_rows(values_only=True):
            row_idx += 1
            if not data_started:
                if row_idx == best_idx:
                    data_started = True
                    # pular a linha de unidades também, se existir
                    continue
                else:
                    continue

            if all(c in (None, "") for c in row):
                continue

            name_idx = col_map.get("name_pt")
            name_val = row[name_idx] if name_idx is not None else None
            if not name_val:
                continue

            name_pt = str(name_val).strip()
            category_pt = None
            if col_map.get("category_pt") is not None:
                cat_val = row[col_map["category_pt"]]
                category_pt = str(cat_val).strip() if cat_val else None

            def num(field):
                idx = col_map.get(field)
                return _parse_float(row[idx]) if idx is not None else None

            writer.writerow({
                "name_pt": name_pt,
                "category_pt": category_pt or "",
                "energy_kcal_100g": num("energy_kcal_100g"),
                "energy_kj_100g": num("energy_kj_100g"),
                "carbohydrates_100g": num("carbohydrates_100g"),
                "proteins_100g": num("proteins_100g"),
                "fat_100g": num("fat_100g"),
                "fiber_100g": num("fiber_100g"),
                "sugars_100g": num("sugars_100g"),
                "sodium_mg_100g": num("sodium_mg_100g"),
                "glycemic_index": num("glycemic_index"),
            })
            written += 1
    return written


def main():
    parser = argparse.ArgumentParser(description="Exporta TACO XLSX para CSV leve (taco_export.csv)")
    parser.add_argument("--xlsx", default=os.path.join(os.path.dirname(__file__), "../../Taco-4a-Edicao.xlsx"), help="Caminho do XLSX de entrada")
    parser.add_argument("--csv", default=os.path.join(os.path.dirname(__file__), "../../taco_export.csv"), help="Caminho do CSV de saída")
    args = parser.parse_args()

    xlsx_path = os.path.abspath(args.xlsx)
    csv_path = os.path.abspath(args.csv)

    if not os.path.exists(xlsx_path):
        print(f"ERRO: XLSX não encontrado: {xlsx_path}")
        sys.exit(1)

    try:
        written = export_csv_from_xlsx(xlsx_path, csv_path)
    except Exception as e:
        print(f"ERRO ao exportar CSV: {e}")
        sys.exit(1)

    print(f"CSV gerado: {csv_path} ({written} linhas)")


if __name__ == "__main__":
    main()