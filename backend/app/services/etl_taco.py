from typing import Dict, List, Optional
import logging
import os
import csv
from unicodedata import normalize
from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.taco_food import TACOFood
from .database import engine

logger = logging.getLogger(__name__)

def _clean_text(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    # Remover espaços extras e normalizar acentos para comparação de headers
    s2 = normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("ASCII")
    return s2.strip().lower()

def _parse_float(val: Optional[str]) -> Optional[float]:
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        # Substituir vírgula por ponto (formato PT-BR)
        v = str(val).replace(",", ".").strip()
        if v == "" or v.lower() in ("na", "nd", "n/a", "-"):
            return None
        return float(v)
    except Exception:
        return None

def _map_headers(headers: List[str]) -> Dict[str, int]:
    """Mapeia headers da TACO para colunas esperadas.

    Suporta tanto os cabeçalhos originais (PT-BR, XLSX) quanto o CSV exportado
    já normalizado (name_pt, energy_kcal_100g, etc.).
    """
    mapping: Dict[str, int] = {}
    normalized = [_clean_text(h) for h in headers]

    def find_exact(column_name: str) -> Optional[int]:
        target = _clean_text(column_name)
        for i, h in enumerate(normalized):
            if h == target:
                return i
        return None

    def find_idx(*keywords) -> Optional[int]:
        for i, h in enumerate(normalized):
            if h is None:
                continue
            if all(k in h for k in keywords):
                return i
        return None

    # name/category
    mapping["name_pt"] = (
        find_exact("name_pt")
        or find_idx("alimento")
        or find_idx("descricao")
        or find_idx("nome")
    )
    mapping["category_pt"] = find_exact("category_pt") or find_idx("grupo") or find_idx("categoria")

    # nutrients
    mapping["energy_kcal_100g"] = find_exact("energy_kcal_100g") or find_idx("energia", "kcal") or find_idx("kcal")
    mapping["energy_kj_100g"] = find_exact("energy_kj_100g") or find_idx("energia", "kj") or find_idx("kj")
    mapping["carbohydrates_100g"] = find_exact("carbohydrates_100g") or find_idx("carbo", "g") or find_idx("carboidratos")
    mapping["proteins_100g"] = find_exact("proteins_100g") or find_idx("proteina")
    mapping["fat_100g"] = find_exact("fat_100g") or find_idx("lipidio") or find_idx("gordura")
    mapping["fiber_100g"] = find_exact("fiber_100g") or find_idx("fibra")
    mapping["sugars_100g"] = find_exact("sugars_100g") or find_idx("acucar") or find_idx("açucar") or find_idx("sacarose")
    mapping["sodium_mg_100g"] = find_exact("sodium_mg_100g") or find_idx("sodio") or find_idx("sódio")
    mapping["glycemic_index"] = find_exact("glycemic_index") or find_idx("indice", "glicemico") or find_idx("ig")

    return mapping

def ingest_taco_excel(path: str) -> Dict[str, int]:
    """Ingestão real da TACO via Excel (XLSX) usando openpyxl.

    - Identifica header automaticamente por linha com >3 colunas preenchidas
    - Faz parsing robusto de números (vírgula/ponto, N/A)
    - Upsert por `name_pt`
    """
    logger.info(f"TACO ingest called for path: {path}")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active  # Assume a primeira planilha

    # Encontrar a linha de header (primeira linha com pelo menos 4 células não vazias)
    header_row_idx = None
    headers: List[str] = []
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) >= 4:
            header_row_idx = row
            headers = [str(c) if c is not None else "" for c in row]
            break
    if not headers:
        raise ValueError("Não foi possível identificar cabeçalhos na planilha TACO")

    col_map = _map_headers(headers)
    if not col_map.get("name_pt"):
        raise ValueError("Coluna de nome do alimento não detectada nos cabeçalhos")

    created = 0
    updated = 0

    # Processar linhas de dados a partir da próxima linha
    data_started = False
    with Session(engine) as session:
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if not data_started:
                # pular até a linha após o header detectado
                if list(row) == list(header_row_idx):
                    data_started = True
                continue

            # Ignorar linhas totalmente vazias
            if all(c in (None, "") for c in row):
                continue

            name_cell_idx = col_map.get("name_pt")
            name_val = row[name_cell_idx] if name_cell_idx is not None else None
            if not name_val:
                continue

            name_pt = str(name_val).strip()
            category_pt = None
            if col_map.get("category_pt") is not None:
                cat_val = row[col_map["category_pt"]]
                category_pt = str(cat_val).strip() if cat_val else None

            # Nutrientes
            energy_kcal_100g = _parse_float(row[col_map["energy_kcal_100g"]]) if col_map.get("energy_kcal_100g") is not None else None
            energy_kj_100g = _parse_float(row[col_map["energy_kj_100g"]]) if col_map.get("energy_kj_100g") is not None else None
            carbohydrates_100g = _parse_float(row[col_map["carbohydrates_100g"]]) if col_map.get("carbohydrates_100g") is not None else None
            proteins_100g = _parse_float(row[col_map["proteins_100g"]]) if col_map.get("proteins_100g") is not None else None
            fat_100g = _parse_float(row[col_map["fat_100g"]]) if col_map.get("fat_100g") is not None else None
            fiber_100g = _parse_float(row[col_map["fiber_100g"]]) if col_map.get("fiber_100g") is not None else None
            sugars_100g = _parse_float(row[col_map["sugars_100g"]]) if col_map.get("sugars_100g") is not None else None
            sodium_mg_100g = _parse_float(row[col_map["sodium_mg_100g"]]) if col_map.get("sodium_mg_100g") is not None else None
            glycemic_index = _parse_float(row[col_map["glycemic_index"]]) if col_map.get("glycemic_index") is not None else None

            # Upsert por name_pt
            stmt = select(TACOFood).where(TACOFood.name_pt == name_pt)
            existing = session.exec(stmt).first()
            if existing:
                existing.category_pt = category_pt
                existing.energy_kcal_100g = energy_kcal_100g
                existing.energy_kj_100g = energy_kj_100g
                existing.carbohydrates_100g = carbohydrates_100g
                existing.proteins_100g = proteins_100g
                existing.fat_100g = fat_100g
                existing.fiber_100g = fiber_100g
                existing.sugars_100g = sugars_100g
                existing.sodium_mg_100g = sodium_mg_100g
                existing.glycemic_index = glycemic_index
                session.add(existing)
                updated += 1
            else:
                item = TACOFood(
                    name_pt=name_pt,
                    category_pt=category_pt,
                    energy_kcal_100g=energy_kcal_100g,
                    energy_kj_100g=energy_kj_100g,
                    carbohydrates_100g=carbohydrates_100g,
                    proteins_100g=proteins_100g,
                    fat_100g=fat_100g,
                    fiber_100g=fiber_100g,
                    sugars_100g=sugars_100g,
                    sodium_mg_100g=sodium_mg_100g,
                    glycemic_index=glycemic_index,
                )
                session.add(item)
                created += 1

            # Commit em lotes para performance
            if (created + updated) % 100 == 0:
                session.commit()

        # Commit final
        session.commit()

    logger.info(f"TACO ingest finished: created={created}, updated={updated}")
    return {"created": created, "updated": updated}


def ingest_taco_csv(path: str) -> Dict[str, int]:
    """Ingestão da TACO via CSV leve (taco_export.csv).

    - Usa DictReader com headers já normalizados (compatíveis com a tabela `taco_foods`)
    - Faz parsing robusto de números (vírgula/ponto, N/A)
    - Upsert por `name_pt`
    """
    logger.info(f"TACO CSV ingest called for path: {path}")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo CSV não encontrado: {path}")

    created = 0
    updated = 0

    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        if not headers or ("name_pt" not in headers):
            raise ValueError("CSV inválido: header 'name_pt' ausente")

        with Session(engine) as session:
            for row in reader:
                name_pt = (row.get("name_pt") or "").strip()
                if not name_pt:
                    continue

                category_pt = (row.get("category_pt") or None)
                energy_kcal_100g = _parse_float(row.get("energy_kcal_100g"))
                energy_kj_100g = _parse_float(row.get("energy_kj_100g"))
                carbohydrates_100g = _parse_float(row.get("carbohydrates_100g"))
                proteins_100g = _parse_float(row.get("proteins_100g"))
                fat_100g = _parse_float(row.get("fat_100g"))
                fiber_100g = _parse_float(row.get("fiber_100g"))
                sugars_100g = _parse_float(row.get("sugars_100g"))
                sodium_mg_100g = _parse_float(row.get("sodium_mg_100g"))
                glycemic_index = _parse_float(row.get("glycemic_index"))

                stmt = select(TACOFood).where(TACOFood.name_pt == name_pt)
                existing = session.exec(stmt).first()
                if existing:
                    existing.category_pt = category_pt
                    existing.energy_kcal_100g = energy_kcal_100g
                    existing.energy_kj_100g = energy_kj_100g
                    existing.carbohydrates_100g = carbohydrates_100g
                    existing.proteins_100g = proteins_100g
                    existing.fat_100g = fat_100g
                    existing.fiber_100g = fiber_100g
                    existing.sugars_100g = sugars_100g
                    existing.sodium_mg_100g = sodium_mg_100g
                    existing.glycemic_index = glycemic_index
                    session.add(existing)
                    updated += 1
                else:
                    item = TACOFood(
                        name_pt=name_pt,
                        category_pt=category_pt,
                        energy_kcal_100g=energy_kcal_100g,
                        energy_kj_100g=energy_kj_100g,
                        carbohydrates_100g=carbohydrates_100g,
                        proteins_100g=proteins_100g,
                        fat_100g=fat_100g,
                        fiber_100g=fiber_100g,
                        sugars_100g=sugars_100g,
                        sodium_mg_100g=sodium_mg_100g,
                        glycemic_index=glycemic_index,
                    )
                    session.add(item)
                    created += 1

                # Commit em lotes para performance
                if (created + updated) % 200 == 0:
                    session.commit()

            # Commit final
            session.commit()

    logger.info(f"TACO CSV ingest finished: created={created}, updated={updated}")
    return {"created": created, "updated": updated}