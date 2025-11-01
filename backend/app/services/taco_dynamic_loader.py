import os
import csv
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from sqlmodel import Session, select

from app.models.taco_food import TACOFood
from .database import engine
from .etl_taco import _map_headers, _parse_float  # reuse header mapping and numeric parsing
from unicodedata import normalize

logger = logging.getLogger(__name__)


def _clean_text(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    s2 = normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("ASCII")
    return s2.strip().lower()


class InMemoryCache:
    """Simple in-memory cache with optional TTL and max size."""

    def __init__(self, ttl_seconds: int = 300, max_items: int = 500):
        self.ttl = ttl_seconds
        self.max_items = max_items
        self.store: Dict[str, Dict[str, Any]] = {}

    def get(self, key: str) -> Optional[Any]:
        entry = self.store.get(key)
        if not entry:
            return None
        if datetime.now() > entry["expires_at"]:
            # expired
            self.store.pop(key, None)
            return None
        return entry["value"]

    def set(self, key: str, value: Any):
        # enforce max size
        if len(self.store) >= self.max_items:
            # remove oldest
            oldest_key = None
            oldest_time = datetime.max
            for k, v in self.store.items():
                if v["expires_at"] < oldest_time:
                    oldest_time = v["expires_at"]
                    oldest_key = k
            if oldest_key:
                self.store.pop(oldest_key, None)

        self.store[key] = {
            "value": value,
            "expires_at": datetime.now() + timedelta(seconds=self.ttl),
        }


class TACODynamicLoader:
    """
    Dynamic loader that searches foods in the root TACO CSV (preferred) or XLSX,
    caches results, and upserts found items into the database.

    Flow:
    - cache -> DB -> CSV/XLSX scan -> upsert -> cache
    """

    def __init__(self, taco_path_env: str = "TACO_FILE_PATH"):
        # Resolve path: prefer env var; else try project root CSV then XLSX
        root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../.."))
        env_path = os.getenv(taco_path_env)
        default_csv = os.path.join(root, "Taco-4a-Edicao.csv")
        default_xlsx = os.path.join(root, "Taco-4a-Edicao.xlsx")

        self.taco_file_path = env_path or (default_csv if os.path.exists(default_csv) else default_xlsx)
        if not os.path.exists(self.taco_file_path):
            logger.warning(
                f"TACO dynamic loader: file not found at '{self.taco_file_path}'."
            )

        self.cache = InMemoryCache(ttl_seconds=600, max_items=1000)

    def _normalize_item(self, row: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "id": "0",
            "source": "taco_dynamic",
            "name": row.get("name_pt"),
            "brands": None,
            "serving_size": None,
            "serving_quantity": None,
            "nutrients_per_100g": {
                "energy_kcal": row.get("energy_kcal_100g"),
                "energy_kj": row.get("energy_kj_100g"),
                "carbohydrates": row.get("carbohydrates_100g"),
                "proteins": row.get("proteins_100g"),
                "fat": row.get("fat_100g"),
                "fiber": row.get("fiber_100g"),
                "sugars": row.get("sugars_100g"),
                "sodium": row.get("sodium_mg_100g"),
                "salt": None,
            },
            "nutriscore": None,
            "ecoscore": None,
            "data_type": "TACO",
            "glycemic_index": row.get("glycemic_index"),
            "category": row.get("category_pt"),
        }

    def _upsert_db_items(self, items: List[Dict[str, Any]]) -> int:
        """Upsert items into DB by name_pt. Returns count of upserted rows."""
        upserted = 0
        with Session(engine) as session:
            for row in items:
                name_pt = row.get("name_pt")
                if not name_pt:
                    continue
                existing = session.exec(select(TACOFood).where(TACOFood.name_pt == name_pt)).first()
                if existing:
                    existing.category_pt = row.get("category_pt")
                    existing.energy_kcal_100g = row.get("energy_kcal_100g")
                    existing.energy_kj_100g = row.get("energy_kj_100g")
                    existing.carbohydrates_100g = row.get("carbohydrates_100g")
                    existing.proteins_100g = row.get("proteins_100g")
                    existing.fat_100g = row.get("fat_100g")
                    existing.fiber_100g = row.get("fiber_100g")
                    existing.sugars_100g = row.get("sugars_100g")
                    existing.sodium_mg_100g = row.get("sodium_mg_100g")
                    existing.glycemic_index = row.get("glycemic_index")
                    session.add(existing)
                else:
                    session.add(TACOFood(
                        name_pt=name_pt,
                        category_pt=row.get("category_pt"),
                        energy_kcal_100g=row.get("energy_kcal_100g"),
                        energy_kj_100g=row.get("energy_kj_100g"),
                        carbohydrates_100g=row.get("carbohydrates_100g"),
                        proteins_100g=row.get("proteins_100g"),
                        fat_100g=row.get("fat_100g"),
                        fiber_100g=row.get("fiber_100g"),
                        sugars_100g=row.get("sugars_100g"),
                        sodium_mg_100g=row.get("sodium_mg_100g"),
                        glycemic_index=row.get("glycemic_index"),
                    ))
                upserted += 1
            session.commit()
        return upserted

    def _scan_csv(self, term: str, page_size: int) -> List[Dict[str, Any]]:
        if not os.path.exists(self.taco_file_path):
            return []
        if not self.taco_file_path.lower().endswith(".csv"):
            return []

        items: List[Dict[str, Any]] = []
        with open(self.taco_file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                return []
            col_map = _map_headers(headers)
            for row in reader:
                if all(c in (None, "") for c in row):
                    continue
                name_idx = col_map.get("name_pt")
                if name_idx is None:
                    continue
                name_val = row[name_idx]
                if not name_val:
                    continue
                name_text = str(name_val).strip()
                if _clean_text(term) not in _clean_text(name_text):
                    continue

                def get_col(key: str) -> Optional[Any]:
                    idx = col_map.get(key)
                    if idx is None:
                        return None
                    return row[idx]

                item = {
                    "name_pt": name_text,
                    "category_pt": (get_col("category_pt") or None),
                    "energy_kcal_100g": _parse_float(get_col("energy_kcal_100g")),
                    "energy_kj_100g": _parse_float(get_col("energy_kj_100g")),
                    "carbohydrates_100g": _parse_float(get_col("carbohydrates_100g")),
                    "proteins_100g": _parse_float(get_col("proteins_100g")),
                    "fat_100g": _parse_float(get_col("fat_100g")),
                    "fiber_100g": _parse_float(get_col("fiber_100g")),
                    "sugars_100g": _parse_float(get_col("sugars_100g")),
                    "sodium_mg_100g": _parse_float(get_col("sodium_mg_100g")),
                    "glycemic_index": _parse_float(get_col("glycemic_index")),
                }
                items.append(item)
                if len(items) >= page_size:
                    break
        return items

    def _scan_xlsx(self, term: str, page_size: int) -> List[Dict[str, Any]]:
        # Use existing ingest mapping helpers to read XLSX incrementally
        try:
            from openpyxl import load_workbook
        except Exception:
            logger.warning("openpyxl not available for XLSX scanning")
            return []

        if not os.path.exists(self.taco_file_path):
            logger.warning("TACO XLSX path does not exist for scan")
            return []
        if not self.taco_file_path.lower().endswith(".xlsx"):
            return []

        items: List[Dict[str, Any]] = []
        wb = load_workbook(self.taco_file_path, data_only=True)
        ws = wb.active

        # Detect header row by index and expected keywords
        header_row_index: Optional[int] = None
        headers: List[str] = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            values = [c for c in row]
            non_empty = [c for c in values if c not in (None, "")]
            if len(non_empty) < 4:
                continue
            candidate_headers = [str(c) if c is not None else "" for c in values]
            col_map_candidate = _map_headers(candidate_headers)
            if col_map_candidate.get("name_pt") is not None:
                header_row_index = i
                headers = candidate_headers
                break
        if not headers or header_row_index is None:
            logger.warning("Could not detect header row in TACO XLSX")
            return []

        col_map = _map_headers(headers)
        name_idx = col_map.get("name_pt")
        if name_idx is None:
            logger.warning("TACO XLSX mapping missing name_pt column")
            return []

        # Iterate data rows from the row after header
        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            if all(c in (None, "") for c in row):
                continue
            name_val = row[name_idx]
            if not name_val:
                continue
            name_text = str(name_val).strip()
            if _clean_text(term) not in _clean_text(name_text):
                continue

            def get_col(key: str) -> Optional[Any]:
                idx = col_map.get(key)
                if idx is None:
                    return None
                return row[idx]

            item = {
                "name_pt": name_text,
                "category_pt": (get_col("category_pt") or None),
                "energy_kcal_100g": _parse_float(get_col("energy_kcal_100g")),
                "energy_kj_100g": _parse_float(get_col("energy_kj_100g")),
                "carbohydrates_100g": _parse_float(get_col("carbohydrates_100g")),
                "proteins_100g": _parse_float(get_col("proteins_100g")),
                "fat_100g": _parse_float(get_col("fat_100g")),
                "fiber_100g": _parse_float(get_col("fiber_100g")),
                "sugars_100g": _parse_float(get_col("sugars_100g")),
                "sodium_mg_100g": _parse_float(get_col("sodium_mg_100g")),
                "glycemic_index": _parse_float(get_col("glycemic_index")),
            }
            items.append(item)
            if len(items) >= page_size:
                break
        return items

    def search(self, term: str, page_size: int = 20) -> Dict[str, Any]:
        """Search through cache, DB, then CSV/XLSX and upsert if needed."""
        key = f"term:{_clean_text(term)}|size:{page_size}"

        # 1) Cache
        cached = self.cache.get(key)
        if cached is not None:
            logger.info(f"TACO dynamic loader: cache hit for term='{term}'")
            return cached

        start = datetime.now()
        items: List[Dict[str, Any]] = []

        # 2) DB first
        with Session(engine) as session:
            results = session.exec(
                select(TACOFood).where(TACOFood.name_pt.ilike(f"%{term}%")).limit(page_size)
            ).all()
            for r in results:
                items.append(self._normalize_item({
                    "name_pt": r.name_pt,
                    "category_pt": r.category_pt,
                    "energy_kcal_100g": r.energy_kcal_100g,
                    "energy_kj_100g": r.energy_kj_100g,
                    "carbohydrates_100g": r.carbohydrates_100g,
                    "proteins_100g": r.proteins_100g,
                    "fat_100g": r.fat_100g,
                    "fiber_100g": r.fiber_100g,
                    "sugars_100g": r.sugars_100g,
                    "sodium_mg_100g": r.sodium_mg_100g,
                    "glycemic_index": r.glycemic_index,
                }))

        # 3) If insufficient, scan CSV/XLSX
        if len(items) < page_size:
            remaining = page_size - len(items)
            scanned = []
            if self.taco_file_path.lower().endswith(".csv"):
                scanned = self._scan_csv(term, remaining)
            elif self.taco_file_path.lower().endswith(".xlsx"):
                scanned = self._scan_xlsx(term, remaining)

            if scanned:
                # Upsert the scanned items into DB
                upsert_count = self._upsert_db_items(scanned)
                logger.info(f"TACO dynamic loader: upserted {upsert_count} items from file")
                # Extend items with normalized scanned entries
                items.extend([self._normalize_item(r) for r in scanned])

        result = {
            "term": term,
            "sources": ["taco_db" if len(items) > 0 else "taco_file"],
            "items": items,
            "total_found": len(items),
            "search_time_ms": round((datetime.now() - start).total_seconds() * 1000, 2),
        }

        # 4) Cache the result for subsequent queries
        self.cache.set(key, result)
        return result